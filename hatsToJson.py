import os
import re
import sys
import copy
import json
import hashlib
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))

IN_FILE = os.path.join(WORKING_DIR, "CustomHatsGM.xlsx")
OUT_FILE = os.path.join(WORKING_DIR, "CustomHats.json")
HAT_DIR = os.path.join(WORKING_DIR, "hats")

def md5(fname):
    hash_md5 = hashlib.md5()
    with open(os.path.join(HAT_DIR, fname), "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()

def cleanString(string):
  if not string: return string
  return string.replace("\r", "").replace("_x000D_", "").replace("\\n", "\n")

def hatsToJson(filename):
  wb = load_workbook(filename, read_only = True)
  
  hatData = {"hats": []}
  for s in wb:
    rows = s.iter_rows(min_col = 1, min_row = 2, max_col = 10, max_row = None)
    headers = []
    for header in s[1]:
      if header.value:
        headers.append(header.value)
        print(header.value)
    
    for row in rows:
      name        = cleanString(row[0].value)
      author      = cleanString(row[1].value)
      package     = cleanString(row[2].value)
      res         = cleanString(row[3].value)
      backres     = cleanString(row[4].value)
      climbres    = cleanString(row[5].value)
      flipres     = cleanString(row[6].value)
      backflipres = cleanString(row[7].value)
      
      if not name or not res:
        continue
      
      options = set(os.path.splitext(res)[0].lower().split("_"))
      
      data = {
        "name": name,
        "author": author,
        "package": package,
        "condition": "None",
        "resource": res,
        "reshasha": md5(res),
      }
      
      if backres:
        data["backresource"] = backres
        data["reshashb"] = md5(backres)
        options.update(os.path.splitext(backres)[0].lower().split("_"))
        
      if climbres:
        data["climbresource"] = climbres
        data["reshashc"] = md5(climbres)
        options.update(os.path.splitext(climbres)[0].lower().split("_"))
        
      if flipres:
        data["flipresource"] = flipres
        data["reshashf"] = md5(flipres)
        options.update(os.path.splitext(flipres)[0].lower().split("_"))
        
      if backflipres:
        data["backflipresource"] = backflipres
        data["reshashbf"] = md5(backflipres)
        options.update(os.path.splitext(backflipres)[0].lower().split("_"))
      
      if "bounce" in options: data["bounce"] = True
      if "adaptive" in options: data["adaptive"] = True
      if "behind" in options: data["behind"] = True
      
      if data:
        hatData["hats"].append(data)
  
  hatData["hats"].sort(key = lambda x: x["name"])
  
  with open(OUT_FILE, "w") as f:
    json.dump(hatData, f, indent=4)

if __name__ == "__main__":
  hatsToJson(IN_FILE)